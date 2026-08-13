"""
Tender Trimesters / Mommies Matter — 90-Day Day-by-Day Marketing Schedule
=======================================================================
Branded schedule with color-coded content types, daily tasks, and status tracking.
"""

import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

import base
base.use_palette_explicit("bottega")  # Dark green — matches Mommies Matter brand

# === Brand color overrides for schedule content types ===
FILL_AFFIRMATION = PatternFill("solid", fgColor="FADADD")  # blush pink
FILL_EDUCATIONAL  = PatternFill("solid", fgColor="CFE3DC")  # sage green
FILL_ENGAGEMENT   = PatternFill("solid", fgColor="FFF3CD")  # butter
FILL_PRODUCT      = PatternFill("solid", fgColor="E8D5C4")  # terracotta light
FILL_EMAIL        = PatternFill("solid", fgColor="D6D4E8")  # lavender
FILL_PARTNERSHIP  = PatternFill("solid", fgColor="D6E4F0")  # soft blue
FILL_LAUNCH       = PatternFill("solid", fgColor="B76E79")  # rose gold
FILL_REST         = PatternFill("solid", fgColor="F5F0E0")  # hearth cream
FILL_PAID         = PatternFill("solid", fgColor="FFE0B2")  # warm amber

FONT_AFFIRMATION = Font(name=base.FONT_NAME, size=11, color="B76E79")
FONT_EDUCATIONAL  = Font(name=base.FONT_NAME, size=11, color="6B7A5A")
FONT_ENGAGEMENT   = Font(name=base.FONT_NAME, size=11, color="8B6914")
FONT_PRODUCT      = Font(name=base.FONT_NAME, size=11, color="C97B5C")
FONT_EMAIL        = Font(name=base.FONT_NAME, size=11, color="6B5B95")
FONT_PARTNERSHIP  = Font(name=base.FONT_NAME, size=11, color="2C5F7C")
FONT_LAUNCH       = Font(name=base.FONT_NAME, size=11, color="FFFFFF", bold=True)
FONT_REST         = Font(name=base.FONT_NAME, size=11, color="8C8A84", italic=True)
FONT_PAID         = Font(name=base.FONT_NAME, size=11, color="D4820A")

CONTENT_TYPE_STYLE = {
    "Affirmation":  (FILL_AFFIRMATION, FONT_AFFIRMATION),
    "Educational":  (FILL_EDUCATIONAL,  FONT_EDUCATIONAL),
    "Engagement":   (FILL_ENGAGEMENT,   FONT_ENGAGEMENT),
    "Product/BTS":  (FILL_PRODUCT,      FONT_PRODUCT),
    "Email":        (FILL_EMAIL,        FONT_EMAIL),
    "Partnership":  (FILL_PARTNERSHIP,  FONT_PARTNERSHIP),
    "Launch Day":   (FILL_LAUNCH,       FONT_LAUNCH),
    "Rest Day":     (FILL_REST,         FONT_REST),
    "Paid Ads":     (FILL_PAID,         FONT_PAID),
}

# === Schedule data ===
# (phase, day_offset, weekday, content_type, platform, task_detail, cta, asset_needed, kpi)
# day_offset: 0 = launch day (Aug 15, 2026)

LAUNCH_DATE = date(2026, 8, 15)

# Phase 1: Pre-Launch (14 days before launch)
PRE_LAUNCH = [
    # Day -14: Saturday Aug 1
    ("Pre-Launch", -14, "Sat", "Product/BTS", "Instagram", "Post first teaser: silhouette of app with 'Something beautiful is growing' text overlay. Use brand gradient moss-to-rose-gold background.", "Save this post — link in bio coming soon", "Teaser image (Canva), brand gradient background", "Impressions > 500, Saves > 50"),
    # Day -13: Sunday Aug 2
    ("Pre-Launch", -13, "Sun", "Rest Day", "—", "Rest and prepare content batch for the week. Review all 30 social captions from the content kit and schedule in Meta Business Suite.", "—", "Content kit docx, Meta Business Suite", "Schedule 5 posts for next week"),
    # Day -12: Monday Aug 3
    ("Pre-Launch", -12, "Mon", "Affirmation", "Instagram Stories + TikTok", "Post affirmation card: 'You are already the mother your child needs.' Show card from affirmation-cards.pdf with soft blush background.", "Share with a mama who needs this today", "Affirmation card image (Week 20 card from PDF)", "Story views > 300, Shares > 20"),
    # Day -11: Tuesday Aug 4
    ("Pre-Launch", -11, "Tue", "Educational", "Instagram Reels + TikTok", "Reel: '3 things nobody tells you about your first trimester' — text-on-screen with voiceover. Mention Tempie AI as the supportive companion.", "Follow for more pregnancy truths — link in bio", "15-sec reel video, teleprompter script", "Views > 1000, Follows > 30"),
    # Day -10: Wednesday Aug 5
    ("Pre-Launch", -10, "Wed", "Engagement", "Instagram", "Carousel: 'What week are you?' — Week 1-40 slide showing baby size comparisons from the app. Ask followers to comment their week.", "Comment your week below — tell us your baby's current size!", "Carousel graphics (10 slides, baby fruit/veg comparisons)", "Comments > 100, Engagement rate > 5%"),
    # Day -9: Thursday Aug 6
    ("Pre-Launch", -9, "Thu", "Product/BTS", "Instagram + Facebook", "BTS post: Show behind-the-scenes of building Tender Trimesters. Helena-Ann's story — why she built this app. Photo of workspace with app on screen.", "This is personal. Link in bio to join the waitlist.", "BTS photo, Helena-Ann headshot, app screenshot", "Link clicks > 50, Waitlist signups > 20"),
    # Day -8: Friday Aug 7
    ("Pre-Launch", -8, "Fri", "Email", "Email (ConvertKit/Mailchimp)", "Send Email #1 of nurture sequence: 'Welcome to the village, mama' — introduce Helena-Ann, share her 'why', soft CTA to waitlist.", "Join the waitlist — be first to try Tender Trimesters free", "Email template in launch-announcement.docx", "Open rate > 40%, CTR > 8%"),
    # Day -7: Saturday Aug 8
    ("Pre-Launch", -7, "Sat", "Product/BTS", "TikTok + Instagram Reels", "Video: Screen recording walkthrough of the app — show HomeScreen with week tracker, Tempie AI chat, calendar view. Add voiceover explaining each feature.", "Waitlist is open — link in bio", "Screen recording (2 min), Loom or OBS, voiceover script", "Views > 2000, Waitlist signups > 30"),
    # Day -6: Sunday Aug 9
    ("Pre-Launch", -6, "Sun", "Rest Day", "—", "Batch create next week's content. Write 5 more captions, design 3 affirmation graphics. Prepare influencer outreach emails from influencer-outreach.docx.", "—", "Canva Pro, content kit, influencer outreach templates", "5 posts ready, 10 outreach emails drafted"),
    # Day -5: Monday Aug 10
    ("Pre-Launch", -5, "Mon", "Affirmation", "Instagram Stories", "Story series (3 slides): 'Your body knows what to do. Trust it.' — Week 28 affirmation card with baby size (eggplant). Final slide: waitlist CTA.", "Join the waitlist — Link in bio", "Affirmation card image, link sticker", "Story taps > 200, Link taps > 30"),
    # Day -4: Tuesday Aug 11
    ("Pre-Launch", -4, "Tue", "Educational", "Instagram + Pinterest", "Infographic carousel: 'Your pregnancy hormones explained' — based on Hormone Horoscope feature. 8 slides covering hCG, progesterone, relaxin, oxytocin.", "Save this for when those hormones hit — and download the app for weekly updates", "Infographic (8 slides), Hormone Horoscope data from app", "Saves > 200, Pins > 50"),
    # Day -3: Wednesday Aug 12
    ("Pre-Launch", -3, "Wed", "Engagement", "Instagram", "Poll sticker: 'What's your biggest pregnancy fear right now?' — 4 options (labor pain, body changes, not bonding, something going wrong). Use results to create content.", "Vote below — we're building something to help with exactly this", "Instagram poll sticker, branded background", "Votes > 200, Profile visits > 50"),
    # Day -2: Thursday Aug 13
    ("Pre-Launch", -2, "Thu", "Email", "Email (ConvertKit/Mailchimp)", "Send Email #2: Helena-Ann's story — 'I was scared too' — personal narrative about pregnancy anxiety, how Tempie was born from that fear. CTA: waitlist.", "Read my full story — and join the waitlist", "Email template in launch-announcement.docx, personal photos", "Open rate > 45%, Replies > 20"),
    # Day -1: Friday Aug 14
    ("Pre-Launch", -1, "Fri", "Product/BTS", "All platforms", "LAUNCH EVE POST: 'Tomorrow, everything changes for mamas.' Countdown graphic. Reveal app icon. Final waitlist push. Share in all pregnancy Facebook groups.", "Waitlist closes at midnight — be there when we launch", "Countdown graphic (Canva), app icon, Facebook group posts", "Waitlist signups > 100 total, Shares > 50"),
]

# Phase 2: Launch Week (Days 0-6)
LAUNCH_WEEK = [
    # Day 0: Saturday Aug 15 — LAUNCH DAY
    ("Launch Week", 0, "Sat", "Launch Day", "All platforms", "LAUNCH DAY: Go live on Vercel. Post launch announcement across all channels. Instagram: 'It's HERE.' with app store link. TikTok: 30-sec hype video. Facebook: full announcement post. Email #3 to waitlist: 'It's live.'", "Download Tender Trimesters FREE — tell every mama you know", "Launch graphics (3 sizes), app store screenshots, email blast template", "500+ app signups Day 1, 1000+ website visits"),
    # Day 1: Sunday Aug 16
    ("Launch Week", 1, "Sun", "Engagement", "Instagram Stories", "Story: 'Show us your screen!' — ask new users to share screenshots of their app home screen (week tracker). Repost the best ones. Give shoutouts.", "Tag us in your screenshot — we're reposting mamas all day", "Story template with 'Show us your screen' text", "User-generated content > 10 posts, Story views > 500"),
    # Day 2: Monday Aug 17
    ("Launch Week", 2, "Mon", "Affirmation", "Instagram + TikTok", "Post Week 12 affirmation: 'You are building a whole human. That is enough.' Tie to first-trimester-checklist.pdf lead magnet — 'Free download in bio.'", "Download your free first trimester survival kit — link in bio", "Affirmation card, first-trimester-checklist.pdf", "Downloads > 100, Follows > 40"),
    # Day 3: Tuesday Aug 18
    ("Launch Week", 3, "Tue", "Educational", "Instagram Reels + TikTok", "Reel: 'How AI is changing pregnancy support' — show Tempie chat in action. Screen record a real conversation where Tempie responds to 'I'm scared of labor.'", "Try Tempie free — link in bio", "Screen recording (45 sec), Tempie chat demo", "Views > 3000, App signups > 50"),
    # Day 4: Wednesday Aug 19
    ("Launch Week", 4, "Wed", "Partnership", "Email + DMs", "Send 10 influencer outreach emails (doulas, pregnancy influencers, mom bloggers). Use templates from influencer-outreach.docx. Follow up with DMs on Instagram.", "Partnership inquiry — let's collaborate", "Influencer outreach templates, partnership pitch one-pager", "10 emails sent, 3 responses"),
    # Day 5: Thursday Aug 20
    ("Launch Week", 5, "Thu", "Product/BTS", "Instagram + Facebook", "Post user testimonials: 'What mamas are saying about Tender Trimesters' — screenshot any early reviews/ratings. If none yet, share Helena-Ann's personal review of using the app.", "Have you tried it yet? Download free — link in bio", "Testimonial graphics, star rating images", "Engagement rate > 6%, Link clicks > 80"),
    # Day 6: Friday Aug 21
    ("Launch Week", 6, "Fri", "Email", "Email (ConvertKit/Mailchimp)", "Send Email #4: Testimonial spotlight — feature a beta user story (or Helena-Ann's own). Introduce Premium features. Soft pitch: $9.99 one-time.", "Unlock everything for $9.99 — or try the free version first", "Email template, premium feature list, pricing graphic", "Open rate > 40%, Premium conversions > 5"),
]

# Phase 3: Growth (Days 7-30)
GROWTH_WEEKS = [
    # Week 2 (Days 7-13)
    ("Week 2: Growth", 7, "Sat", "Affirmation", "Instagram Stories", "Saturday affirmation series — post 3 story slides with handwritten-style affirmation from the app. Use Dancing Script font on blush background.", "Share with a mama who needs this", "3 affirmation graphics (Stories size)", "Story views > 400"),
    ("Week 2: Growth", 8, "Sun", "Rest Day", "—", "Content batching day: Write next 7 days of captions. Design 5 new graphics. Review analytics from Week 1 — what performed best? Double down on that format.", "—", "Analytics dashboard, Canva, caption template", "7 posts scheduled, analytics reviewed"),
    ("Week 2: Growth", 9, "Mon", "Educational", "Instagram + Pinterest", "Carousel: 'Fear to Flame — how to reframe pregnancy anxiety' — walkthrough of the app feature. Explain the 3 stages: Ember, Spark, Flame.", "This feature changed how I process fear — try it free", "Feature walkthrough graphics (6 slides)", "Saves > 150, App signups > 30"),
    ("Week 2: Growth", 10, "Tue", "Engagement", "Instagram", "'Fill in the blank: Before I got pregnant, I wish someone told me ___' — engage comments, reply to every single one personally. This builds community.", "Drop your answer below — let's talk about it", "Branded engagement graphic", "Comments > 150, Reply rate > 80%"),
    ("Week 2: Growth", 11, "Wed", "Product/BTS", "TikTok + Reels", "Video: 'What's inside the Name Garden?' — show the name suggestion themes, the 'plant a seed' feature, chosen name celebration. Fun and light.", "Find your baby's name in the app — free", "Screen recording (60 sec), name garden demo", "Views > 2000, App opens > 40"),
    ("Week 2: Growth", 12, "Thu", "Email", "Email", "Send Email #5: 'Your baby wrote you a letter' — showcase Letters from Baby feature. Include a sample AI-generated letter. CTA: generate your first letter.", "Read your baby's letter — it's free in the app", "Sample letter screenshot, email template", "Open rate > 35%, App opens > 60"),
    ("Week 2: Growth", 13, "Fri", "Paid Ads", "Meta Ads (Facebook + Instagram)", "Launch first paid campaign: $10/day budget. Target: pregnant women 22-38, interest in pregnancy apps, baby names, meditation. Creative: testimonial-style video.", "Download free — Tender Trimesters", "Ad creative (video), Meta Ads Manager setup", "CPA < $3, 50+ installs from ads"),
    # Week 3 (Days 14-20)
    ("Week 3: Momentum", 14, "Sat", "Affirmation", "TikTok + Instagram", "TikTok trend: 'Tell me you're pregnant without telling me you're pregnant' — use app screenshots as reveal. Pair with trending audio.", "Tag a pregnant mama — she needs this app", "TikTok video (15 sec), trending audio", "Views > 5000, Follows > 100"),
    ("Week 3: Momentum", 15, "Sun", "Rest Day", "—", "Plan Week 4 content. Review paid ad performance — adjust targeting if CPA > $5. Reach out to 5 more influencers. Follow up with previous outreach.", "—", "Ad analytics, influencer tracker, content calendar", "Ads optimized, 5 follow-ups sent"),
    ("Week 3: Momentum", 16, "Mon", "Educational", "Instagram + Blog", "Deep-dive post: 'Understanding your pregnancy dreams' — tie to DreamKeeper feature. Explain common pregnancy dream themes (water, birth, animals). CTA: start logging dreams.", "Your dreams mean something — start capturing them", "Long-form caption (200+ words), dream theme graphics", "Saves > 200, DreamKeeper feature opens > 50"),
    ("Week 3: Momentum", 17, "Tue", "Engagement", "Instagram", "'This or That: Pregnancy Edition' — 10 slides of pregnancy choices (find out gender vs surprise, epidural vs natural, etc.). High engagement format.", "Save your results and tag a pregnant friend", "This or That carousel (10 slides)", "Engagement rate > 7%, Shares > 80"),
    ("Week 3: Momentum", 18, "Wed", "Partnership", "Email + DMs", "Follow up with all influencers who haven't responded. Send partnership pitch to 5 local doulas/midwives. Offer free Premium in exchange for honest review.", "Free Premium access for honest review", "Partnership follow-up templates, Premium access codes", "5 new partnerships initiated"),
    ("Week 3: Momentum", 19, "Thu", "Product/BTS", "Instagram + TikTok", "Feature spotlight: Memory Capsule — 'Seal a memory for your child to open someday.' Show the create flow, unlock options (1yr, 5yr, 18th birthday). Emotional appeal.", "What would you seal for your baby? Try it free", "Feature demo video (45 sec), memory capsule screenshots", "Views > 2500, Feature opens > 60"),
    ("Week 3: Momentum", 20, "Fri", "Email", "Email", "Weekly digest email: 'This week in your pregnancy' — dynamic content based on user's week. Include affirmation, tip, and feature spotlight.", "Open the app for your full weekly update", "Email template, dynamic content logic", "Open rate > 35%, App opens > 100"),
    # Week 4 (Days 21-27)
    ("Week 4: Scale", 21, "Sat", "Affirmation", "Instagram Stories", "Weekend affirmation: 'You don't have to do this perfectly. You just have to do it.' Story poll: 'What's your self-care today?' with options.", "Tell us your self-care below", "Affirmation graphic, poll sticker", "Story taps > 500, Poll votes > 200"),
    ("Week 4: Scale", 22, "Sun", "Rest Day", "—", "Monthly review: Pull all analytics. Calculate: total signups, conversion rate, top content, best platform, ad ROAS. Adjust month 2 strategy.", "—", "Analytics from all platforms, spreadsheet for tracking", "Monthly report completed"),
    ("Week 4: Scale", 23, "Mon", "Educational", "Pinterest + Instagram", "Pin 10 pregnancy-related infographics to Pinterest (hormone horoscope, fear stages, dream meanings). Each pin links to blog or app landing page.", "Follow us on Pinterest for weekly pregnancy guides", "10 Pinterest graphics (1000x1500px), pin descriptions with keywords", "Pinterest impressions > 5000, Clicks > 100"),
    ("Week 4: Scale", 24, "Tue", "Engagement", "Instagram", "'Ask me anything about pregnancy' — go live on Instagram for 20 min. Answer questions about the app, pregnancy tips, share personal stories. Save to Highlights.", "Join us live — ask anything", "IG Live setup, promo story 1hr before", "Live viewers > 50, Story replay views > 300"),
    ("Week 4: Scale", 25, "Wed", "Product/BTS", "TikTok + Reels", "Video: 'Build your birth playlist in the app' — show Birth Playlist feature, play song suggestions for each phase (early labor, active labor, pushing, golden hour).", "What songs are on your birth playlist? Tell us below", "Birth playlist demo (60 sec), song list graphic", "Views > 3000, Comments > 100"),
    ("Week 4: Scale", 26, "Thu", "Paid Ads", "Meta Ads", "Scale winning ad creative. Create 2 new ad variations based on best-performing organic post. Increase budget to $15/day if CPA < $3.", "Download Tender Trimesters free", "New ad creatives (2 variations), Meta Ads Manager", "CPA < $2.50 (improved), 80+ installs from ads"),
    ("Week 4: Scale", 27, "Fri", "Email", "Email", "Month 1 wrap-up: 'What a month, mamas' — share stats (users joined, letters generated, fears reframed). Feature 2-3 user testimonials. Premium upsell.", "You've made this first month incredible — here's what's next", "Email template, user testimonials, stats graphic", "Open rate > 35%, Premium conversions > 3%"),
    # Remaining growth days (28-29)
    ("Week 4: Scale", 28, "Sat", "Affirmation", "Instagram + TikTok", "Affirmation video: Helena-Ann reads an affirmation from the app in her own voice. Soft, warm, personal. 'You are enough, mama.' Loopable for TikTok.", "Follow for daily affirmations for your pregnancy journey", "Video of Helena-Ann (30 sec), mic setup, quiet space", "Views > 4000, Follows > 80"),
    ("Week 4: Scale", 29, "Sun", "Rest Day", "—", "Plan Month 2 content calendar. Identify top 5 performing posts and recreate with variations. Draft 2 new lead magnets (second trimester checklist, fear journal).", "—", "Analytics, content calendar template, Canva", "Month 2 plan ready, 2 lead magnets outlined"),
]

# Phase 4: Sustain (Days 30-59) — summary rows, repeating weekly patterns
SUSTAIN_WEEKS = []
weekly_patterns = [
    ("Mon", "Affirmation", "Instagram Stories + TikTok", "Weekly affirmation post — rotate through app affirmation cards. Share on Stories with interactive poll sticker.", "Share with a mama who needs this today"),
    ("Tue", "Educational", "Instagram + Pinterest", "Educational carousel or reel — pregnancy tips, feature deep-dives, hormone insights, dream analysis.", "Save this and share with your due-date group"),
    ("Wed", "Engagement", "Instagram", "Interactive content — polls, quizzes, this-or-that, Q&A, fill-in-the-blank. Reply to every comment.", "Comment below — we read and reply to every single one"),
    ("Thu", "Product/BTS", "TikTok + Instagram Reels", "Feature spotlight video — screen record one premium feature. Show real usage. Emotional storytelling.", "Try this feature free — link in bio"),
    ("Fri", "Email", "Email", "Weekly digest or nurture email — rotate between: weekly update, feature spotlight, user testimonial, Helena-Ann's letter.", "Open the app for your full weekly update"),
    ("Sat", "Affirmation + Paid Ads", "Instagram + Meta Ads", "Affirmation post + manage paid ads. Review ad performance, adjust budgets, create new creatives as needed.", "Follow for daily pregnancy affirmations"),
    ("Sun", "Rest Day", "—", "Content batching, analytics review, influencer follow-ups, strategy planning for next week.", "—"),
]

for week_num in range(5, 9):  # Weeks 5-8
    phase_name = f"Week {week_num}: Sustain"
    base_day = 28 + (week_num - 5) * 7
    for i, (wday, ctype, platform, task, cta) in enumerate(weekly_patterns):
        day_offset = base_day + i
        asset = "Recycled from content bank" if "Rest" not in ctype else "—"
        kpi = "Engagement rate > 5%" if "Engagement" in ctype else "—"
        if "Paid" in ctype:
            kpi = "CPA < $2.50, 100+ weekly installs from ads"
            asset = "Ad creatives, Meta Ads Manager"
        if "Email" in ctype and "Rest" not in ctype:
            kpi = "Open rate > 30%, App opens > 80"
            asset = "Email template"
        SUSTAIN_WEEKS.append((phase_name, day_offset, wday, ctype, platform, task, cta, asset, kpi))

# Phase 5: Expand (Days 60-89)
EXPAND_WEEKS = []
expand_patterns = [
    ("Mon", "Partnership", "Email + DMs", "Week of partnership outreach — contact 5 new influencers, doulas, or birth educators. Send review copies of Premium.", "Free Premium for honest review — DM us"),
    ("Tue", "Educational", "All platforms", "Long-form content: blog post or video essay on a pregnancy topic. Repurpose into 5+ micro-content pieces for social.", "Read the full guide — link in bio"),
    ("Wed", "Engagement", "Instagram", "Community-building content — user spotlight, 'mama of the week,' partner feature, bump photo showcase.", "Tag us to be featured — we celebrate every mama"),
    ("Thu", "Product/BTS", "TikTok + Reels", "New feature announcement or update showcase. If no new features, show a 'day in the life' using the app.", "Update your app for the latest features"),
    ("Fri", "Email", "Email", "Monthly newsletter: top tips, user stories, new features, premium offer. Segment: free users vs premium.", "See what's new this month — and unlock premium"),
    ("Sat", "Affirmation + Paid Ads", "Instagram + Meta Ads", "Affirmation + paid ads management. A/B test new ad creatives. Scale budget if ROAS > 3x.", "Follow for daily affirmations"),
    ("Sun", "Rest Day", "—", "Strategic review: update 90-day plan, set new KPIs, plan next quarter's content themes and partnership goals.", "—"),
]

for week_num in range(9, 13):  # Weeks 9-12
    phase_name = f"Week {week_num}: Expand"
    base_day = 56 + (week_num - 9) * 7
    for i, (wday, ctype, platform, task, cta) in enumerate(expand_patterns):
        day_offset = base_day + i
        asset = "Partnership templates" if "Partnership" in ctype else "Content repurposed from blog/video"
        kpi = "5+ new partnerships/week" if "Partnership" in ctype else "—"
        if "Paid" in ctype:
            kpi = "ROAS > 3x, Scale to $25/day if profitable"
            asset = "New ad creatives (A/B test), Meta Ads Manager"
        if "Email" in ctype and "Rest" not in ctype:
            kpi = "Open rate > 30%, Premium upsell > 2%"
            asset = "Monthly newsletter template"
        if "Rest" in ctype:
            asset = "—"
        EXPAND_WEEKS.append((phase_name, day_offset, wday, ctype, platform, task, cta, asset, kpi))

# === Build workbook ===
ALL_DATA = PRE_LAUNCH + LAUNCH_WEEK + GROWTH_WEEKS + SUSTAIN_WEEKS + EXPAND_WEEKS

wb = Workbook()

# === Sheet 1: Day-by-Day Schedule ===
ws = wb.active
ws.title = "90-Day Marketing Schedule"

# Column widths
ws.column_dimensions["A"].width = 3   # margin
col_widths = {"B": 16, "C": 12, "D": 11, "E": 16, "F": 20, "G": 55, "H": 40, "I": 35, "J": 32, "K": 10}
for col_letter, w in col_widths.items():
    ws.column_dimensions[col_letter].width = w

# Title
ws.merge_cells("B2:J2")
title_cell = ws.cell(row=2, column=2, value="Tender Trimesters — 90-Day Day-by-Day Marketing Schedule")
title_cell.font = Font(name=base.FONT_NAME, size=16, bold=True, color=base.PRIMARY)
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 32

# Subtitle
ws.merge_cells("B3:J3")
sub_cell = ws.cell(row=3, column=2, value="Mommies Matter  |  Launch: August 15, 2026  |  Helena-Ann Baker")
sub_cell.font = base.font_caption()
ws.row_dimensions[3].height = 20

# Headers
headers = ["Phase", "Date", "Day", "Content Type", "Platform", "Task Detail", "CTA / Message", "Asset Needed", "KPI Target", "Status"]
header_row = 4
for col_idx, h in enumerate(headers, 2):
    cell = ws.cell(row=header_row, column=col_idx, value=h)
    cell.fill = base.fill_header()
    cell.font = base.font_header()
    cell.alignment = base.align_header()
    cell.border = base.border_header()
ws.row_dimensions[header_row].height = 28

# Data rows
data_start = 5
for row_idx, (phase, day_off, wday, ctype, platform, task, cta, asset, kpi) in enumerate(ALL_DATA):
    row = data_start + row_idx
    actual_date = LAUNCH_DATE + timedelta(days=day_off)

    values = [phase, actual_date, wday, ctype, platform, task, cta, asset, kpi, "To Do"]
    for col_idx, val in enumerate(values, 2):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = base.font_body()
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Apply content type color to the type column (E) and whole row background
        if col_idx == 5:  # Content Type column
            if ctype in CONTENT_TYPE_STYLE:
                fill, font = CONTENT_TYPE_STYLE[ctype]
                cell.fill = fill
                cell.font = font
        elif col_idx == 12:  # Status column
            cell.font = base.font_body()
            cell.alignment = base.align_header()

    # Date column formatting
    ws.cell(row=row, column=3).number_format = "MMM D, YYYY"
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")  # Day

    ws.row_dimensions[row].height = 42

# Freeze panes
ws.freeze_panes = "B5"

# Data validation for Status column
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"To Do,In Progress,Done,Skipped"', allow_blank=True)
dv.error = "Please select a valid status"
dv.errorTitle = "Invalid status"
ws.add_data_validation(dv)
for row in range(data_start, data_start + len(ALL_DATA)):
    dv.add(ws.cell(row=row, column=11))

# === Sheet 2: Content Type Legend ===
ws2 = wb.create_sheet("Content Type Legend")
ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 50
ws2.column_dimensions["D"].width = 30

ws2.merge_cells("B2:D2")
ws2.cell(row=2, column=2, value="Content Type Color Legend").font = Font(name=base.FONT_NAME, size=16, bold=True, color=base.PRIMARY)
ws2.row_dimensions[2].height = 32

legend_headers = ["Content Type", "Description", "When to Use"]
for col_idx, h in enumerate(legend_headers, 2):
    cell = ws2.cell(row=4, column=col_idx, value=h)
    cell.fill = base.fill_header()
    cell.font = base.font_header()
    cell.alignment = base.align_header()
    cell.border = base.border_header()

legend_data = [
    ("Affirmation", "Inspirational quotes, affirmation cards, emotional support content", "Monday & Saturday — start and end the week with warmth"),
    ("Educational", "Pregnancy tips, feature deep-dives, infographics, hormone insights", "Tuesday — mid-week value content that gets saved"),
    ("Engagement", "Polls, quizzes, Q&A, this-or-that, fill-in-the-blank, user spotlights", "Wednesday — peak engagement day for comments"),
    ("Product/BTS", "App features, behind-the-scenes, screen recordings, testimonials", "Thursday — showcase what makes Tender Trimesters special"),
    ("Email", "Nurture sequences, weekly digests, feature announcements, newsletters", "Friday — end-of-week inbox presence"),
    ("Partnership", "Influencer outreach, doula/midwife partnerships, collaboration pitches", "Week 3+ — once initial content foundation is set"),
    ("Launch Day", "All-hands marketing push across every channel simultaneously", "Day 0 only — maximum visibility"),
    ("Paid Ads", "Meta Ads management, A/B testing, budget scaling, ROAS optimization", "Week 2+ — after organic content proves what resonates"),
    ("Rest Day", "Content batching, analytics review, strategy planning, self-care", "Every Sunday — prevent burnout, plan ahead"),
]

for i, (ctype, desc, when) in enumerate(legend_data):
    row = 5 + i
    ws2.cell(row=row, column=2, value=ctype)
    ws2.cell(row=row, column=3, value=desc)
    ws2.cell(row=row, column=4, value=when)

    if ctype in CONTENT_TYPE_STYLE:
        fill, font = CONTENT_TYPE_STYLE[ctype]
        ws2.cell(row=row, column=2).fill = fill
        ws2.cell(row=row, column=2).font = font

    for col in range(2, 5):
        ws2.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
    ws2.row_dimensions[row].height = 30

ws2.freeze_panes = "B5"

# === Sheet 3: Weekly KPI Tracker ===
ws3 = wb.create_sheet("Weekly KPI Tracker")
ws3.column_dimensions["A"].width = 3
for col_letter, w in {"B": 14, "C": 12, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14}.items():
    ws3.column_dimensions[col_letter].width = w

ws3.merge_cells("B2:K2")
ws3.cell(row=2, column=2, value="Weekly KPI Tracker — Track Your Growth").font = Font(name=base.FONT_NAME, size=16, bold=True, color=base.PRIMARY)
ws3.row_dimensions[2].height = 32

kpi_headers = ["Week", "New Signups", "Premium Conv.", "Instagram Foll.", "Email Subs", "Ad Spend", "Ad Installs", "CPA", "Top Platform", "Notes"]
for col_idx, h in enumerate(kpi_headers, 2):
    cell = ws3.cell(row=4, column=col_idx, value=h)
    cell.fill = base.fill_header()
    cell.font = base.font_header()
    cell.alignment = base.align_header()
    cell.border = base.border_header()

for week_i in range(13):
    row = 5 + week_i
    week_label = f"Week {week_i + 1}" if week_i < 4 else f"Week {week_i + 1}"
    ws3.cell(row=row, column=2, value=week_label)
    # Leave other cells empty for user to fill
    for col in range(3, 12):
        ws3.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[row].height = 22

ws3.freeze_panes = "C5"

# === Save ===
OUTPUT = "/home/z/my-project/download/marketing-schedule-90day.xlsx"
wb.save(OUTPUT)
print(f"Saved to {OUTPUT}")
print(f"Total schedule rows: {len(ALL_DATA)}")
